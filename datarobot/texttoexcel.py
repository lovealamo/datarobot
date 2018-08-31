# coding=utf-8
import requests,time
import re
import xlrd
from openpyxl import workbook  # 写入Excel表所用
from openpyxl import load_workbook  # 读取Excel表所用
import numpy as np
from bs4 import BeautifulSoup as bs
import os
import sys
from openpyxl.styles import colors
from openpyxl.styles import numbers
from openpyxl.styles import Font, Color
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection

numchainofeachmachine = 3;
columnwidth = 12.0
summarycolumnwidth=13.0

fillgreen = PatternFill(fill_type="solid", start_color=colors.GREEN,end_color=colors.DARKGREEN)
fillyellow = PatternFill(fill_type="solid", start_color=colors.YELLOW,end_color=colors.DARKYELLOW)
fillred = PatternFill(fill_type="solid", start_color=colors.RED,end_color=colors.DARKRED)
fillblack = PatternFill(fill_type="solid", start_color=colors.BLACK,end_color=colors.BLACK)

def loadData(fileName):
    testData=[]
    resultdata=[]
    intlist = []
    print(fileName)
    with open(fileName) as txtData:
        lines=txtData.readlines()
        #print(lines)
        for line in lines:
            temp = re.findall('^[1-9]+',line)
            if temp :
                testData.append(line)
        
        for i in range(len(testData)):
            if (i % 6) == 0:
                resultdata.append(testData[i].split())
        #print('capture chain1,2,3 set final freq=',resultdata)
        #convert str integeter into int integer.比如把下面这个list转成下下面的intlist
        #[['365','355','325'],['315','325','300'],['200','310','325']]
        #[[365,355,325],[315,325,300],[200,310,325]]
        for var in resultdata:
            templist = [int(i) for i in var]
            #print(templist)
            intlist.append(templist)
        #print('intlist = ',intlist)
        return intlist

      
def getalllogfile(file_dir):
    for root, dirs, files in os.walk(file_dir):
        #print(root) #当前目录路径  
        #print(dirs) #当前路径下所有子目录  
        #print(files) #当前路径下所有非目录子文件
        return (dirs,files)

def generateOneSheet(index,totallist,wb,resultdir):
    #获取http://10.76.6.11/mms/miner/matrix.do下载下来的机器的datasheet的数组
    successscannum =0
    failscannum = 0

    sheetrecord=[0,0,0,0,0,0,0,0,0,0,0,0,0]
    #runtime id on 0
    sheetrecord[0]=(index+1);
    #pcb version on 2
    sheetrecord[2]='V3.76';
    #firmwareversion on 3
    sheetrecord[3]='201808271014'
    #scan style on 4
    sheetrecord[4]='按列顺序扫频'



    rb = xlrd.open_workbook('./data/'+resultdir+'.xls')
    table = rb.sheet_by_name("T11-V3.76(A7V4BIN2)2")
    #table.cell(2,3).value
    calupower=[]
    staticlist=[]

    avgpower=[]
    calpower=[]
    avgcalrate=[]

    for var in range(1,21):
    #for var in range(1,int(argv[2])):
        #方便计算算力比率
        temp=[]
        temp.append(table.cell(var,2).value)
        temp.append(table.cell(var,11).value)
        print("temp=",temp[0],temp[1])
        calupower.append(temp)
        
        #纯为了方便后期统计平均值，标准差，算力比率方差

        if table.cell(var,2).value and table.cell(var,11).value:
            avgpower.append(table.cell(var,2).value)
            calpower.append(table.cell(var,11).value)
            avgcalrate.append(table.cell(var,2).value/table.cell(var,11).value)

    staticlist.append(avgpower)
    staticlist.append(calpower)
    staticlist.append(avgcalrate)

    #ws = wb.active  # 获取当前正在操作的表对象
    ws = wb.create_sheet(title = resultdir)

    row0=['IP Address', '算力板', '第一列频率','第二列频率','第三列频率','第四列频率', '第五列频率',
                '第六列频率', '第七列频率', '第八列频率', '第九列频率', '第十列频率','实测算力','理论算力', '算力百分比']
    ws.append(row0)
    
    #log元素从第二行开始，第一行是上面的列标。2，3，4一个元素，5，6，7一个元素。
    linenum=2
    # 往表中写入标题行,以列表形式写入,True 返回文件列标，False返回目录列表
    datalogpath = './data/'+resultdir+'/'

    #alllogfile装的是参数目录下面的所有非目录的文件
    (folder,alllogfile)=getalllogfile(datalogpath)
    print("alllogfile _________________________________________length =",len(alllogfile))
    print(alllogfile)
    for index in range(0,len(alllogfile)):
        #用_切开文件名和ip,templist[0]文件，templist[1]:ip
        templist=alllogfile[index].split('_')
        print("temp________________",templist)

        #统计ip范围
        if index == 0:#第一个ip
            recordip = templist[1]+'~'
        elif index == (len(alllogfile) - 1):#最后一个ip
            recordip = recordip+templist[1]
            sheetrecord[1] = recordip

        resultlist = loadData(datalogpath+alllogfile[index])
        if len(resultlist):
            successscannum+=1
            for i in range(len(resultlist)):
                print("this is i=",i)
                itemlist = []
                #添加IP进入行内容itemlist
                itemlist.append(templist[1])
                #添加chain x信息进入templist
                itemlist.append('chain'+str(i+1))
                #itemlist.append(resultlist[i])
                for value in range(len(resultlist[i])):
                    itemlist.append(resultlist[i][value])
                #仅在第一个chain的信息打印实测算力和理论算力
                if i == 0:
                    #实际算力
                    itemlist.append(calupower[index][0])
                    #理论算力
                    itemlist.append(calupower[index][1])
                    if int(calupower[index][1]) != 0:
                        print("int(calupower[index][1])=",int(calupower[index][1]))
                        itemlist.append(calupower[index][0]/calupower[index][1])
                ws.append(itemlist)
        else:
            print("read log file didn't have frequency data\n")
            failscannum+=1
            ws.append([templist[1],'chain1',0,0,0,0,0,0,0,0,0,0,0,0,0.0])
            ws.append([templist[1],'chain2',0,0,0,0,0,0,0,0,0,0,0,0,0.0])
            ws.append([templist[1],'chain3',0,0,0,0,0,0,0,0,0,0,0,0,0.0])

        #合并必要的单元格
        #print('index=',index)
        newlinenum = linenum+numchainofeachmachine
        #print('linenum=',linenum,'nextlinenum=',newlinenum)
        mergecellgrp = 'A'+str(linenum)+':'+'A'+str(newlinenum-1)
        #print('mergecellgrp=',mergecellgrp)
        ws.merge_cells(mergecellgrp)
        mergecellgrp='M'+str(linenum)+':'+'M'+str(newlinenum-1)
        ws.merge_cells(mergecellgrp)
        mergecellgrp='N'+str(linenum)+':'+'N'+str(newlinenum-1)
        ws.merge_cells(mergecellgrp)
        mergecellgrp='O'+str(linenum)+':'+'O'+str(newlinenum-1)
        ws.merge_cells(mergecellgrp)
        linenum = newlinenum
    
    #设置最后一列为百分比并且根据值设置演示以示区分。
    setfillvalue = fillgreen
    for cell in ws[chr(ord('A')+ws.max_column-1)]:
        cell.number_format = '0.00%'
        #cell.width = 40.0
        if cell.row != 1:
            if isinstance(cell.value,float):
                if cell.value >= 0.98:
                    cell.fill = fillgreen
                    setfillvalue = fillgreen
                elif ((cell.value) >= 0.9) and ((cell.value) < 0.98):
                    cell.fill = fillyellow
                    setfillvalue = fillyellow
                elif cell.value < 0.9 and cell.value >= 0.1:
                    cell.fill = fillred
                    setfillvalue = fillred
                elif cell.value <0.1:
                    cell.fill = fillblack
                    setfillvalue = fillblack
            else:
                cell.fill = setfillvalue
    
    #对频率数据进行颜色区分
    square = 'C2:'+'L'+str(ws.max_row)
    for row in ws.iter_rows(square):
        for cell in row:
            if cell.value >360:
                cell.fill = fillred
            elif cell.value >=300 and cell.value <= 360:
                cell.fill = fillgreen
            elif cell.value >= 200 and cell.value<300:
                cell.fill = fillyellow
            elif cell.value <200 and cell.value >0:
                cell.fill = fillblack
    
    #调整列宽
    for var in range(ord('A'),(ord('A')+ws.max_column)):
        #print('width=',ws[chr(var)])
        ws.column_dimensions[chr(var)].width = columnwidth

    #scan failed number on 5
    sheetrecord[5]=(failscannum)
    #scan successfull on 6
    sheetrecord[6]=(successscannum)

    #actual power avg on 7
    sheetrecord[7]=(np.mean(staticlist[0]))

    #calculate power avg on 8
    sheetrecord[8]=(np.mean(staticlist[1]))

    #avg rate on 9
    sheetrecord[9]=(np.mean(staticlist[0])/np.mean(staticlist[1]))

    #rate mean on 10
    sheetrecord[10]=(np.mean(staticlist[2]))

    #std on 11
    sheetrecord[11]=(np.std(staticlist[2]))

    #var on 12
    sheetrecord[12]=(np.var(staticlist[2]))

    #add one sheet report into total data list
    totallist.append(sheetrecord)

if __name__ == '__main__':
    #读取存在的Excel表测试
    #wb = load_workbook('test.xlsx') #加载存在的Excel表
    #a_sheet = wb.get_sheet_by_name('Sheet1') #根据表名获取表对象
    #for row in a_sheet.rows: #遍历输出行数据
    #for cell in row: #每行的每一个单元格
    #print cell.value,
    #创建Excel表并写入数据
    #ws.write(5,2,"string")
    totallist=[]
    '''
    argv = sys.argv
    if argv == None:
        print("请输入下载从http://10.76.6.11/mms/miner/matrix.do下来的excel统计文件。")
        exit()
    print(argv)
    
    if os.path.exists('autoexcel.xlsx'):
        print("autoexcel.xlsx exists")
        wb = load_workbook('autoexcel.xlsx')
    else:
        print("autoexcel.xlsx no exists")
        wb = workbook.Workbook()
    '''
    
    if os.path.exists('autoexcel.xlsx'):
        print("autoexcel.xlsx exists")
        os.remove('autoexcel.xlsx')
    wb = workbook.Workbook()

    sheetnames = wb.sheetnames
    print('sheetnames=',sheetnames,'len(sheetname)=',len(sheetnames))
    if (len(sheetnames) <= 1) and sheetnames[0] != 'Summay':
        print('one sheet and rename')
        wstotal = wb.active
        wstotal.title='Summary'
        wstotal.append(['Runtime次数','ip范围','PCB版本','软件版本','扫频方式','扫频失败数','有效统计机器数量','实际算力平均值','理论算力平均值','平均算力比率','算力比率平均值','算力比率标准差','算力比率方差'])
        #ws = wb.create_sheet(title = 'runtime1')
    else:
        print('new sheets')
        wstotal = wb['total']
        #runtime=len(sheetnames)
        #ws= wb.create_sheet(title=('runtime'+str(runtime)))
    
    (dirs,files)= getalllogfile("./data/")
    #print('dirs=',dirs)
    #print('files=',files)
    #wb.save('autoexcel.xlsx')
    #exit()

    for var in range(0,len(dirs)):
        #xls与要分析的目录同名所以这里只用了dirs名字。
        generateOneSheet(var,totallist,wb,dirs[var])
    
    
    for var in range(0,len(totallist)):
        wstotal.append(totallist[var])



    #设置单元格为百分比
    for column in ['J','K']:
        for cell in wstotal[column]:
            cell.number_format = '0.00%'
            if cell.row != 1:
                if cell.value >= 0.98:
                    cell.fill = fillgreen
                elif ((cell.value) >= 0.9) and ((cell.value) < 0.98):
                    cell.fill = fillyellow
                elif cell.value < 0.9 and cell.value >= 0.1:
                    cell.fill = fillred
                elif cell.value <0.1:
                    cell.fill = fillblack
  
    #调整summary列宽
    for var in range(ord('A'),(ord('A')+wstotal.max_column)):
        #print('width=',ws[chr(var)])
        wstotal.column_dimensions[chr(var)].width = summarycolumnwidth
    wb.save('autoexcel.xlsx')